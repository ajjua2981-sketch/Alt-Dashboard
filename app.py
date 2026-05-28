"""
Alternate Drug Dashboard
Run with:  streamlit run app.py
"""

import io
import pandas as pd
import streamlit as st

from api_client import lookup_alternate_drugs
from config import API_CONFIG, EXPORT_CONFIG
from log_parser import parse_log

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Alternate Drug Dashboard",
    page_icon="💊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
.badge-pass { color:#fff; background:#28a745; padding:3px 12px; border-radius:4px; font-size:0.82rem; font-weight:600; }
.badge-fail { color:#fff; background:#dc3545; padding:3px 12px; border-radius:4px; font-size:0.82rem; font-weight:600; }
.badge-noalt { color:#212529; background:#ffc107; padding:3px 12px; border-radius:4px; font-size:0.82rem; font-weight:600; }
.badge-err  { color:#fff; background:#6c757d; padding:3px 12px; border-radius:4px; font-size:0.82rem; font-weight:600; }
div[data-testid="metric-container"] { background:#f8f9fa; border-radius:8px; padding:12px; }
.stDataFrame { font-size: 0.85rem; }
</style>
""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.title("💊 Alternate Drug")
    st.subheader("Dashboard")
    st.divider()

    st.markdown("**API Configuration**")
    st.code(
        f"Endpoint : {API_CONFIG['base_url']}{API_CONFIG['endpoint']}\n"
        f"Batch Size: {API_CONFIG['batch_size']} NDCs per call\n"
        f"Timeout  : {API_CONFIG['timeout_seconds']}s",
        language="text",
    )

    st.divider()
    st.markdown("**Instructions**")
    st.markdown("""
1. Download the Excel template below
2. Paste your log entries into the **logTXT** column (one per row)
3. Upload the completed file
4. Click **Look Up & Compare**
5. View Pass / Fail results and export
""")

    st.markdown("**Log message format detected:**")
    st.code(
        "Requested NDC: <NDC> Substituted with\n"
        "Alternate NDC: <ALT_NDC> For DAW Code: <DAW>\n"
        "Drug Source: <SRC> Substitution Indicator: <SI>",
        language="text",
    )

    st.divider()

    # ── Template download ─────────────────────────────────────────────────────
    st.markdown("**Download Input Template**")
    template_df = pd.DataFrame(columns=["Reference ID", "Case ID", "logTXT"])
    template_buffer = io.BytesIO()
    with pd.ExcelWriter(template_buffer, engine="openpyxl") as writer:
        template_df.to_excel(writer, index=False, sheet_name="Input")
        worksheet = writer.sheets["Input"]
        worksheet.column_dimensions["A"].width = 20
        worksheet.column_dimensions["B"].width = 20
        worksheet.column_dimensions["C"].width = 80
        from openpyxl.styles import Font, PatternFill
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    st.download_button(
        label="⬇️ Download Excel Template",
        data=template_buffer.getvalue(),
        file_name="log_input_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

# ── Main area ─────────────────────────────────────────────────────────────────
st.title("💊 Alternate Drug Dashboard")
st.caption(
    "Upload an Excel file with a **logTXT** column. "
    "Each row is a log entry — the dashboard extracts NDC fields, "
    "calls the API, and compares the API result against the log's expected alternate NDC."
)

st.divider()

# ── File upload ───────────────────────────────────────────────────────────────
uploaded_file = st.file_uploader(
    "Upload Excel file (.xlsx)",
    type=["xlsx"],
    help="Excel file must have columns: 'Reference ID', 'Case ID', 'logTXT'",
)

if uploaded_file:
    # ── Parse Excel ───────────────────────────────────────────────────────────
    try:
        xl         = pd.ExcelFile(uploaded_file, engine="openpyxl")
        sheet_name = xl.sheet_names[0]
        if len(xl.sheet_names) > 1:
            sheet_name = st.selectbox(
                "Multiple sheets found — select the input sheet:",
                options=xl.sheet_names,
            )
        df_input = xl.parse(sheet_name)
        df_input.columns = df_input.columns.str.strip()
    except Exception as exc:
        st.error(f"Could not read Excel file: {exc}")
        st.stop()

    # ── Validate column ───────────────────────────────────────────────────────
    if "logTXT" not in df_input.columns:
        st.error(
            f"Excel file must have a **logTXT** column.\n\n"
            f"Found columns: {', '.join(df_input.columns.tolist())}\n\n"
            "Download the template from the sidebar to get the correct format."
        )
        st.stop()

    # Carry Reference ID and Case ID if present, defaulting to empty string
    for col in ["Reference ID", "Case ID"]:
        if col not in df_input.columns:
            df_input[col] = ""

    df_input = df_input[["Reference ID", "Case ID", "logTXT"]].copy()
    df_input["Reference ID"] = df_input["Reference ID"].astype(str).str.strip().replace("nan", "")
    df_input["Case ID"]      = df_input["Case ID"].astype(str).str.strip().replace("nan", "")
    df_input["logTXT"]       = df_input["logTXT"].astype(str).str.strip()
    df_input = df_input[df_input["logTXT"].str.len() > 0]
    df_input = df_input[df_input["logTXT"] != "nan"].reset_index(drop=True)

    if len(df_input) == 0:
        st.warning("The uploaded file has no data rows.")
        st.stop()

    # ── Parse log entries ─────────────────────────────────────────────────────
    parsed_rows = []
    for _, row in df_input.iterrows():
        result = parse_log(row["logTXT"])
        base = {"reference_id": row["Reference ID"], "case_id": row["Case ID"], "logTXT": row["logTXT"]}
        if result:
            parsed_rows.append({**result, **base, "parse_ok": True})
        else:
            parsed_rows.append({
                "requested_ndc": "", "log_alternate_ndc": "",
                "daw_code": "", "drug_source": "", "substitution_indicator": "",
                **base, "parse_ok": False,
            })

    df_parsed = pd.DataFrame(parsed_rows)
    total       = len(df_parsed)
    parse_ok    = df_parsed["parse_ok"].sum()
    parse_fail  = total - parse_ok

    # ── Preview ───────────────────────────────────────────────────────────────
    st.subheader("📂 Parsed Log Preview")

    pm1, pm2, pm3 = st.columns(3)
    pm1.metric("Total Rows", total)
    pm2.metric("✅ Parsed OK", int(parse_ok))
    pm3.metric("⚠️ Parse Failed", int(parse_fail))

    if parse_fail > 0:
        st.warning(
            f"{int(parse_fail)} row(s) did not contain the expected substitution message "
            "and will be skipped during API lookup."
        )

    preview_df = df_parsed[df_parsed["parse_ok"]].reset_index(drop=True)[[
        "reference_id", "case_id", "requested_ndc", "log_alternate_ndc", "daw_code", "drug_source", "substitution_indicator"
    ]].rename(columns={
        "reference_id":           "Reference ID",
        "case_id":                "Case ID",
        "requested_ndc":          "Requested NDC",
        "log_alternate_ndc":      "Log: Expected Alt NDC",
        "daw_code":               "DAW Code",
        "drug_source":            "Drug Source",
        "substitution_indicator": "Substitution Indicator",
    })

    st.dataframe(preview_df, use_container_width=True, height=220)

    st.divider()

    # ── Lookup button ─────────────────────────────────────────────────────────
    if st.button("🔍 Look Up & Compare", type="primary", use_container_width=True):
        valid_rows = df_parsed[df_parsed["parse_ok"]].reset_index(drop=True)

        progress_bar = st.progress(0, text="Calling API...")
        status_text  = st.empty()
        all_results  = []
        total_valid  = len(valid_rows)

        for i, row in valid_rows.iterrows():
            status_text.text(f"Processing row {i + 1} of {total_valid}...")

            drug = [{
                "ndc":                   row["requested_ndc"],
                "dawCode":               row["daw_code"],
                "substitutionIndicator": row["substitution_indicator"],
            }]

            results, errors = lookup_alternate_drugs(drug)

            if errors:
                all_results.append({
                    "Reference ID":           row["reference_id"],
                    "Case ID":                row["case_id"],
                    "Requested NDC":          row["requested_ndc"],
                    "DAW Code":               row["daw_code"],
                    "Drug Source":            row["drug_source"],
                    "Substitution Indicator": row["substitution_indicator"],
                    "Log: Expected Alt NDC":  row["log_alternate_ndc"],
                    "API: Alt NDC":           "",
                    "API: Alt Drug Name":     "",
                    "Result":                 "API Error",
                    "Error Detail":           errors[0].get("error", ""),
                })
            elif results:
                api_alt_ndc  = results[0].get("Alternate Drug NDC", "")
                api_alt_name = results[0].get("Alternate Drug Name", "")
                log_alt_ndc  = row["log_alternate_ndc"]

                if not api_alt_ndc:
                    verdict = "No Alternate"
                elif api_alt_ndc == log_alt_ndc:
                    verdict = "Pass"
                else:
                    verdict = "Fail"

                all_results.append({
                    "Reference ID":           row["reference_id"],
                    "Case ID":                row["case_id"],
                    "Requested NDC":          row["requested_ndc"],
                    "DAW Code":               row["daw_code"],
                    "Drug Source":            row["drug_source"],
                    "Substitution Indicator": row["substitution_indicator"],
                    "Log: Expected Alt NDC":  log_alt_ndc,
                    "API: Alt NDC":           api_alt_ndc,
                    "API: Alt Drug Name":     api_alt_name,
                    "Result":                 verdict,
                    "Error Detail":           "",
                })

            progress_bar.progress((i + 1) / total_valid)

        progress_bar.empty()
        status_text.empty()

        st.session_state["results"]      = all_results
        st.session_state["parse_fail"]   = int(parse_fail)
        st.session_state["total"]        = total

# ── Results section ───────────────────────────────────────────────────────────
if "results" in st.session_state and st.session_state["results"]:
    results = st.session_state["results"]
    df_results = pd.DataFrame(results)

    st.divider()
    st.subheader("📊 Results Summary")

    r_pass   = (df_results["Result"] == "Pass").sum()
    r_fail   = (df_results["Result"] == "Fail").sum()
    r_noalt  = (df_results["Result"] == "No Alternate").sum()
    r_err    = (df_results["Result"] == "API Error").sum()
    r_pf     = st.session_state.get("parse_fail", 0)

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("✅ Pass",          int(r_pass))
    m2.metric("❌ Fail",          int(r_fail))
    m3.metric("⚠️ No Alternate",  int(r_noalt))
    m4.metric("🔴 API Error",     int(r_err))
    m5.metric("⛔ Parse Failed",  int(r_pf))

    st.divider()

    # ── Filter ────────────────────────────────────────────────────────────────
    st.subheader("🔎 Filter Results")
    f1, f2 = st.columns([2, 3])
    with f1:
        all_verdicts = sorted(df_results["Result"].unique().tolist())
        filter_result = st.multiselect("Result", options=all_verdicts, default=all_verdicts)
    with f2:
        filter_ndc = st.text_input("Search by Requested NDC", placeholder="leave blank for all")

    df_view = df_results.copy()
    if filter_result:
        df_view = df_view[df_view["Result"].isin(filter_result)]
    if filter_ndc.strip():
        df_view = df_view[
            df_view["Requested NDC"].str.contains(filter_ndc.strip(), case=False, na=False)
        ]

    st.divider()

    # ── Results table ─────────────────────────────────────────────────────────
    st.subheader(f"📋 Comparison Results ({len(df_view)} records)")

    def _verdict_badge(val):
        badge_map = {
            "Pass":         "badge-pass",
            "Fail":         "badge-fail",
            "No Alternate": "badge-noalt",
            "API Error":    "badge-err",
        }
        css = badge_map.get(str(val), "badge-err")
        return f'<span class="{css}">{val}</span>'

    if df_view.empty:
        st.info("No results match the current filter.")
    else:
        df_display = df_view.copy()
        df_display["Result"] = df_display["Result"].apply(_verdict_badge)
        cols_order = [
            "Reference ID", "Case ID",
            "Requested NDC", "DAW Code", "Drug Source", "Substitution Indicator",
            "Log: Expected Alt NDC", "API: Alt NDC", "API: Alt Drug Name", "Result",
        ]
        # Only show Error Detail if there are API errors
        if (df_view["Result"].str.contains("API Error")).any() or r_err > 0:
            cols_order.append("Error Detail")
        df_display = df_display[[c for c in cols_order if c in df_display.columns]]
        st.write(df_display.to_html(escape=False, index=False), unsafe_allow_html=True)

    st.divider()

    # ── Export ────────────────────────────────────────────────────────────────
    st.subheader("📥 Export Results")
    fname = EXPORT_CONFIG["default_filename"]
    ex1, ex2 = st.columns(2)

    with ex1:
        csv_data = df_view.to_csv(index=False).encode("utf-8")
        st.download_button(
            label="⬇️ Download as CSV",
            data=csv_data,
            file_name=f"{fname}.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with ex2:
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            df_view.to_excel(writer, index=False, sheet_name="Results")
            ws = writer.sheets["Results"]
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 26
            from openpyxl.styles import Font, PatternFill
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        st.download_button(
            label="⬇️ Download as Excel",
            data=excel_buffer.getvalue(),
            file_name=f"{fname}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
